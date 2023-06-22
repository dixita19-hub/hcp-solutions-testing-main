import json
import os
import typing
import time
import random
import csv

import polling
import structlog
import re
import openpyxl
from openpyxl import utils
from datetime import datetime, timedelta

from pathlib import Path

from openpyxl.reader.excel import load_workbook
from _pytest.fixtures import FixtureRequest

from pytest_bdd import parsers, given, when, then
from pytest_check import check

from bp_core.backend.common.step_definitions.steps_common import set_request_endpoint, set_request_headers, \
    add_json_payload, make_api_request, get_api_response, setup_basic_auth
from bp_core.frontend.common.helpers.app import context_manager
from bp_core.frontend.common.utils.analytics_parser import Analytics
from bp_core.frontend.common.utils.containers import WindowSize
from selenium.common.exceptions import NoSuchElementException, TimeoutException

from assertpy import assert_that
from bp_core.frontend.common.helpers.selenium_generics import SeleniumGenerics
from bp_core.frontend.common.utils.locator_parser import Locators
from bp_core.frontend.common.helpers.standalone_compare import are_two_images_look_same
from bp_core.utils.faker_data import DataUtils
from bp_core.utils.gherkin_utils import (data_table_horizontal_converter, data_table_vertical_converter)
from bp_core.frontend.common.helpers.web_compare import (are_two_webelements_look_same, are_two_webpages_look_same)

from bp_core.utils import env_variables
from bp_core.utils.bs_storage import BPStorage

logger = structlog.get_logger(__name__)
MOBILE_SUFFIX = "_mob"
UI_API_CALL = "ui_api_call"

"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

ENVIRONMENT

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@given(parsers.re("I set the locale for locators to '(?P<locale>.*)'"),
    converters=dict(locale=env_variables.env_formatted), )
@given(parsers.re("I set the language for locators to '(?P<locale>.*)'"),
    converters=dict(locale=env_variables.env_formatted), )
@when(parsers.re("I set the locale for locators to '(?P<locale>.*)'"),
    converters=dict(locale=env_variables.env_formatted), )
@when(parsers.re("I set the language for locators to '(?P<locale>.*)'"),
    converters=dict(locale=env_variables.env_formatted), )
def set_locale(locators: Locators, locale):
    # TODO: Check Why is setting a locale a step def? Shouldn't this be a CLI argument / config value?
    locators.locale = locale


# WEB context Predefined Step
@given(parsers.re("I pause for '(?P<seconds>.*)' s"), converters=dict(seconds=int))
@when(parsers.re("I pause for '(?P<seconds>.*)' s"), converters=dict(seconds=int))
def pause_execution(seconds: int):
    time.sleep(seconds)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("I get text from element '(?P<locator_path>.*)' and save as environment variable '(?P<env_var>.*)'"))
@when(parsers.re("I get text from element '(?P<locator_path>.*)' and save as environment variable '(?P<env_var>.*)'"))
def get_text_from_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str, env_var: str, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    os.environ[env_var] = returned_text


@given(parsers.re("I get text from element '(?P<locator_path>.*)' between '(?P<initial_string>.*)' and '(?P<final_string>.*)' boundaries, and save as environment variable '(?P<env_var>.*)'"),
    converters=dict(initial_string=env_variables.env_formatted, final_string=env_variables.env_formatted), )
@when(parsers.re("I get text from element '(?P<locator_path>.*)' between '(?P<initial_string>.*)' and '(?P<final_string>.*)' boundaries, and save as environment variable '(?P<env_var>.*)'"),
    converters=dict(initial_string=env_variables.env_formatted, final_string=env_variables.env_formatted), )
def get_text_from_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str, initial_string: str, final_string:str, env_var: str, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        returned_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    os.environ[env_var] = returned_text[returned_text.index(initial_string) + len(initial_string): returned_text.index(final_string)]


# WEB context Predefined Step
@given(parsers.re("I write within the HTML report the environment variable '(?P<env_var>.*)' value"))
@when(parsers.re("I write within the HTML report the environment variable '(?P<env_var>.*)' value"))
@then(parsers.re("I write within the HTML report the environment variable '(?P<env_var>.*)' value"))
def write_html_report_os_environ_value(selenium_generics: SeleniumGenerics, env_var: str):
    if BPStorage.get_env_vars_for_html() is not None:
        BPStorage.save_env_vars_for_html({**BPStorage.get_env_vars_for_html(), **{env_var: os.environ.get(env_var)}})
    else:
        BPStorage.save_env_vars_for_html({**{env_var: os.environ.get(env_var)}})


# WEB context Predefined Step
@given(parsers.re("I store '(?P<key>.*)' environment variable in .local.env config file"))
@when(parsers.re("I store '(?P<key>.*)' environment variable in .local.env config file"))
def store_env_variable_in_local_env(key: str):
    local_config_file = Path.cwd() / "configs" / ".local.env"
    if local_config_file.is_file():
        from dotenv import set_key
        if os.environ.get(key, None):
            value = os.environ[key]
            set_key(local_config_file.as_posix(), key, value)
        else:
            raise KeyError(f"Environment variable: {key} does not exits")
    else:
        raise FileNotFoundError("File not found: .local.env")


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

BROWSER / DEVICE

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


@given(parsers.re("I '(?P<cache_option>disable|enable)' the cache"))
@when(parsers.re("I '(?P<cache_option>disable|enable)' the cache"))
def disable_cache(driver, cache_option: str):
    cache_disabled_value = True if cache_option == 'disable' else False
    driver.execute_cdp_cmd("Network.setCacheDisabled", {"cacheDisabled": cache_disabled_value})


# WEB context Predefined Step
@given(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' contains the value '(?P<value>.*)'"),
        converters=dict(value=env_variables.env_formatted))
@when(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' contains the value '(?P<value>.*)'"),
        converters=dict(value=env_variables.env_formatted))
def check_cookie_content(selenium_generics: SeleniumGenerics, soft_assert: str, name, value: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(selenium_generics.cookie_value(name)).contains(value)
    else:
        assert_that(selenium_generics.cookie_value(name)).contains(value)


# WEB context Predefined Step
@given(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' does not contain the value '(?P<value>.*)'"),
        converters=dict(value=env_variables.env_formatted))
@when(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' does not contain the value '(?P<value>.*)'"),
        converters=dict(value=env_variables.env_formatted))
def check_cookie_content_is_not(selenium_generics: SeleniumGenerics, soft_assert: str, name, value: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(selenium_generics.cookie_value(name)).does_not_contain(value)
    else:
        assert_that(selenium_generics.cookie_value(name)).does_not_contain(value)


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' exists"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' exists"))
def check_cookie_exists(selenium_generics: SeleniumGenerics, soft_assert: str, name):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(selenium_generics.cookie_name(name)).is_equal_to(name)
    else:
        assert_that(selenium_generics.cookie_name(name)).is_equal_to(name)


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' does not exist"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The cookie '(?P<name>.*)' does not exist"))
def check_cookie_does_not_exist(selenium_generics: SeleniumGenerics, soft_assert: str, name):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.get_cookie(name) is None
    else:
        assert selenium_generics.get_cookie(name) is None


# WEB context Predefined Step
@given("I fetch existing cookies from the site")
@when("I fetch existing cookies from the site")
def fetch_cookies(selenium_generics: SeleniumGenerics):
    selenium_generics.get_all_cookies()


# WEB context Predefined Step
@given(
    parsers.re("I update the value of newly added cookie '(?P<name>.*)' with '(?P<value>.*)' for path '(?P<path>.*)'"),
        converters=dict(value=env_variables.env_formatted))
@given(parsers.re("I set the cookie '(?P<name>.*)' with value '(?P<value>.*)' for path '(?P<path>.*)'"),
        converters=dict(value=env_variables.env_formatted))
@when(
    parsers.re("I update the value of newly added cookie '(?P<name>.*)' with '(?P<value>.*)' for path '(?P<path>.*)'"),
        converters=dict(value=env_variables.env_formatted))
@when(parsers.re("I set the cookie '(?P<name>.*)' with value '(?P<value>.*)' for path '(?P<path>.*)'"),
        converters=dict(value=env_variables.env_formatted))
def check_cookie_content(selenium_generics: SeleniumGenerics, name, value: str, path="/"):
    selenium_generics.add_cookie(name, value, path=path)


# WEB context Predefined Step
@given(parsers.re("I delete the cookie '(?P<name>.*)'"))
@when(parsers.re("I delete the cookie '(?P<name>.*)'"))
def delete_cookie(selenium_generics: SeleniumGenerics, name):
    selenium_generics.delete_cookie(name)


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect the cookies to be present"))
def check_cookies_presence(selenium_generics: SeleniumGenerics, soft_assert: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.get_all_cookies()
    else:
        assert selenium_generics.get_all_cookies()


# WEB context Predefined Step
@then(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?I expect cookie '(?P<name>.*)' with value '(?P<value>.*)' to be present"),
        converters=dict(value=env_variables.env_formatted))
@then(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?I expect the value of newly added cookie '(?P<name>.*)' to be updated with '(?P<value>.*)'"),
        converters=dict(value=env_variables.env_formatted))
def check_cookie_presence(selenium_generics: SeleniumGenerics, soft_assert: str, name, value: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.cookie_value(name) == value
    else:
        assert selenium_generics.cookie_value(name) == value


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect cookie '(?P<name>.*)' to be deleted"))
def check_cookie_delete(selenium_generics: SeleniumGenerics, soft_assert: str, name):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.get_cookie(name) is None
    else:
        assert selenium_generics.get_cookie(name) is None


# WEB context Predefined Step
@given(parsers.re("The browser resolution is '(?P<width>.*)' per '(?P<height>.*)'"),
    converters=dict(width=env_variables.env_formatted, height=env_variables.env_formatted), )
@given(parsers.re("My screen resolution is '(?P<width>.*)' by '(?P<height>.*)' pixels"),
    converters=dict(width=env_variables.env_formatted, height=env_variables.env_formatted), )
@when(parsers.re("The browser resolution is '(?P<width>.*)' per '(?P<height>.*)'"),
    converters=dict(width=env_variables.env_formatted, height=env_variables.env_formatted), )
@when(parsers.re("My screen resolution is '(?P<width>.*)' by '(?P<height>.*)' pixels"),
    converters=dict(width=env_variables.env_formatted, height=env_variables.env_formatted), )
def window_size(selenium_generics: SeleniumGenerics, width: int, height: int):
    selenium_generics.set_window_size(WindowSize(int(width), int(height)))


# WEB context Predefined Step
@given("Browser is maximized")
@when("Browser is maximized")
def maximize(selenium_generics: SeleniumGenerics):
    selenium_generics.maximize_window()


# WEB context Predefined Step
@given(parsers.re("There is just one (browser tab|window) open"))
@when(parsers.re("There is just one (browser tab|window) open"))
def close_all_but_first_tab(selenium_generics: SeleniumGenerics):
    windows = selenium_generics.window_handles
    windows_to_close = windows[1:]
    while windows_to_close:
        selenium_generics.switch_to_known_window(windows_to_close.pop())
        selenium_generics.close_active_window()


# WEB context Predefined Step
@given(parsers.re("I open new tab with url '(?P<page_url>.*)'"),
    converters=dict(page_url=env_variables.env_formatted),)
@when(parsers.re("I open new tab with url '(?P<page_url>.*)'"),
    converters=dict(page_url=env_variables.env_formatted),)
def open_specific_tab(selenium_generics: SeleniumGenerics, base_url: str, page_url: str):
    selenium_generics.open_new_tab(f"{base_url}{page_url}")


# WEB context Predefined Step
@given(parsers.re("I switch to iframe '(?P<locator_path>.*)'"))
@when(parsers.re("I switch to iframe '(?P<locator_path>.*)'"))
def switch_to_iframe(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    selenium_generics.switch_context_to_iframe(locators.parse_and_get(locator_path, selenium_generics))


# WEB context Predefined Step
@given(parsers.re("I switch back from iframe'"))
@when(parsers.re("I switch back from iframe"))
def switch_back_from_iframe(selenium_generics: SeleniumGenerics):
    selenium_generics.switch_context_to_default_content()


# WEB context Predefined Step
@given(parsers.re("I close the last opened window"))
@given(parsers.re("I close the last opened tab"))
@when(parsers.re("I close the last opened window"))
@when(parsers.re("I close the last opened tab"))
def close_last_opened_window(selenium_generics: SeleniumGenerics):
    selenium_generics.close_last_window()


# WEB context Predefined Step
@given(parsers.re("I close the current opened tab"))
@when(parsers.re("I close the current opened tab"))
def close_current_opened_tab(selenium_generics: SeleniumGenerics):
    selenium_generics.close_active_window()


# WEB context Predefined Step
@given(parsers.re("I close the tab with (url|number) '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted),)
@when(parsers.re("I close the tab with (url|number) '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted),)
def close_specific_tab(selenium_generics: SeleniumGenerics, base_url: str, value: typing.Union[str, int]):
    if value.isdigit():
        selenium_generics.close_window_by_position(int(value))
    else:
        selenium_generics.close_window_by_url(f"{base_url}{value}")


# WEB context Predefined Step
@given(parsers.re("I focus the last opened window"))
@given(parsers.re("I focus the last opened tab"))
@when(parsers.re("I focus the last opened window"))
@when(parsers.re("I focus the last opened tab"))
def switch_to_last(selenium_generics: SeleniumGenerics):
    selenium_generics.switch_to_last_window()


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?There are '(?P<count>.*)' (tabs|windows) currently opened"),
    converters=dict(count=env_variables.env_formatted))
def check_number_of_tabs(selenium_generics: SeleniumGenerics, soft_assert: str, count: int):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.num_of_windows == int(count)
    else:
        assert selenium_generics.num_of_windows == int(count)


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A new (tab|window) is opened"))
def check_new_window(selenium_generics: SeleniumGenerics, soft_assert: str):
    # todo: refactor this method in future releases as appropriate
    # error case: how does this verify the scenario where there is already more than one tab opened prior to this step.
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.num_of_windows > 1
    else:
        assert selenium_generics.num_of_windows > 1


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A new (tab|window) is not opened"))
def check_no_new_window(selenium_generics: SeleniumGenerics, soft_assert: str):
    # todo: refactor this method in future releases as appropriate
    # error case: how does this verify the scenario where there is already more than one tab opened prior to this step.
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.num_of_windows == 1
    else:
        assert selenium_generics.num_of_windows == 1


# WEB context Predefined Step
@given(parsers.re("I take a screenshot"))
@when(parsers.re("I take a screenshot"))
@then(parsers.re("I take a screenshot"))
def take_a_screenshot(selenium_generics: SeleniumGenerics):
    # Declaration of the BDD step that is used to take a screenshot during the scenario
    ...


# MOBILE Predefined Step
@when(parsers.re("I turn '(?P<network_status>ON|OFF)' the network connectivity in Browserstack"))
@then(parsers.re("I turn '(?P<network_status>ON|OFF)' the network connectivity in Browserstack"))
def change_network_status(selenium, request: FixtureRequest, network_status: str):
    NETWORK_STATUS = "CHANGE_NETWORK_STATUS_IN_BROWSERSTACK"
    payload = {"networkProfile": "no-network"} if network_status.lower() == "off" else {"networkProfile": "reset"}
    api_response = dict()
    set_request_endpoint(request, base_url="https://api-cloud.browserstack.com/app-automate/sessions/",
                         endpoint=f"{selenium.session_id}/update_network.json", request_name=NETWORK_STATUS)
    selenium_host = request.session.config.option.selenium_host
    BS_USERNAME = selenium_host[0: selenium_host.index(":")]
    BS_PASSWORD = selenium_host[selenium_host.index(":") + 1: selenium_host.index("@")]
    setup_basic_auth(request=request, username=BS_USERNAME, password=BS_PASSWORD)
    add_json_payload(request, json_payload=payload, request_name=NETWORK_STATUS)
    make_api_request(request, api_response, request_type="PUT", request_name=NETWORK_STATUS)
    assert get_api_response(request, api_response, request_name=NETWORK_STATUS).status_code == 200


# MOBILE Predefined Step
@when("I reset the mobile app")
@then("I reset the mobile app")
def reset_app(selenium_generics: SeleniumGenerics):
    current_context = selenium_generics.get_current_context()
    selenium_generics.reset()
    if selenium_generics.is_android():
        selenium_generics.switch_context(current_context)


# MOBILE Predefined Step
@when(parsers.re("I put the mobile app in background for '(?P<seconds>.*)' seconds"),
    converters=dict(seconds=env_variables.env_formatted))
@then(parsers.re("I put the mobile app in background for '(?P<seconds>.*)' seconds"),
    converters=dict(seconds=env_variables.env_formatted))
def background_app(selenium_generics: SeleniumGenerics, seconds: int):
    selenium_generics.background_app(seconds=int(seconds))


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

PAGE & URL

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@given(parsers.re("I am on the (url|page|site) '(?P<page_url>.*)'"),
    converters=dict(page_url=env_variables.env_formatted), )
@when(parsers.re("I am on the (url|page|site) '(?P<page_url>.*)'"),
    converters=dict(page_url=env_variables.env_formatted), )
def open_webpage(selenium_generics: SeleniumGenerics, base_url: str, page_url: str):
    selenium_generics.navigate_to_url(f"{base_url}{page_url}")


# WEB context Predefined Step
@given(parsers.re("I get current browser url and store it to '(?P<env_var>.*)'"))
@when(parsers.re("I get current browser url and store it to '(?P<env_var>.*)'"))
@then(parsers.re("I get current browser url and store it to '(?P<env_var>.*)'"))
def store_url_env_var(selenium_generics, env_var: str):
    os.environ[env_var] = selenium_generics.current_url


@given(parsers.re("I switch to tab with (url|number) '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("I switch to tab with (url|number) '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def switch_specific_tab(selenium_generics: SeleniumGenerics, base_url: str, value: typing.Union[str, int]):
    if value.isdigit():
        selenium_generics.switch_to_window_by_position(int(value))
    else:
        selenium_generics.switch_tab_by_url(f"{base_url}{value}")


# WEB context Predefined Step
@given(parsers.re("I navigate to external page '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
@when(parsers.re("I navigate to external page '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
def navigate_to_external_page(selenium_generics: SeleniumGenerics, url: str):
    selenium_generics.navigate_to_url(url)


# WEB context Predefined Step
@given("I navigate back to the previous page")
@when("I navigate back to the previous page")
@then("I navigate back to the previous page")
def navigate_to_previous_page(selenium_generics: SeleniumGenerics):
    selenium_generics.one_page_backward()


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The title is '(?P<title>.*)'"),
    converters=dict(title=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The title is '(?P<title>.*)'"),
    converters=dict(title=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that the title is '(?P<title>.*)'"),
    converters=dict(title=env_variables.env_formatted), )
def page_title(selenium_generics: SeleniumGenerics, soft_assert: str, title: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.does_title_equals(title), f"Title Mismatch: {selenium_generics.title} vs {title}"
    else:
        assert selenium_generics.does_title_equals(title), f"Title Mismatch: {selenium_generics.title} vs {title}"


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The title is not '(?P<title>.*)'"),
    converters=dict(title=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The title is not '(?P<title>.*)'"),
    converters=dict(title=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that the title is not '(?P<title>.*)'"),
    converters=dict(title=env_variables.env_formatted), )
def page_title_is_not(selenium_generics: SeleniumGenerics, soft_assert: str, title: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert not selenium_generics.does_title_equals(
                title), f"Title Mismatch: {selenium_generics.title} vs {title}"
    else:
        assert not selenium_generics.does_title_equals(title), f"Title Mismatch: {selenium_generics.title} vs {title}"


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that the title contains '(?P<title>.*)'"),
    converters=dict(title=env_variables.env_formatted), )
def check_title_contains(selenium_generics: SeleniumGenerics, soft_assert: str, title: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.does_title_contains(title), f"Mismatch: {title} not in {selenium_generics.title}"
    else:
        assert selenium_generics.does_title_contains(title), f"Mismatch: {title} not in {selenium_generics.title}"


# WEB context Predefined Step
@then(
    parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that the title does not contain '(?P<title>.*)'"),
        converters=dict(title=env_variables.env_formatted), )
def check_title_not_contains(selenium_generics: SeleniumGenerics, soft_assert: str, title: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert not selenium_generics.does_title_contains(title), f"Mismatch: {title} in {selenium_generics.title}"
    else:
        assert not selenium_generics.does_title_contains(title), f"Mismatch: {title} in {selenium_generics.title}"


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page url is '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page url is '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page url is '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
def given_page_url_is(selenium_generics: SeleniumGenerics, soft_assert: str, url: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.does_url_equals(url), f"URL mismatch: {selenium_generics.current_url} vs {url}"
    else:
        assert selenium_generics.does_url_equals(url), f"URL mismatch: {selenium_generics.current_url} vs {url}"


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page url is not '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page url is not '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page url is not '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
def given_page_url_is_not(selenium_generics: SeleniumGenerics, soft_assert: str, url: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.does_url_not_equals(url), f"URL mismatch: {selenium_generics.current_url} vs {url}"
    else:
        assert selenium_generics.does_url_not_equals(url), f"URL mismatch: {selenium_generics.current_url} vs {url}"


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page (path is|url contains) '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
def check_page_url_contains(selenium_generics: SeleniumGenerics, soft_assert: str, url: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.does_current_url_contains(url)
    else:
        assert selenium_generics.does_current_url_contains(url)


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The page (path is not|url does not contain) '(?P<url>.*)'"),
    converters=dict(url=env_variables.env_formatted), )
def check_page_url_not_contains(selenium_generics: SeleniumGenerics, soft_assert: str, url: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert not selenium_generics.does_current_url_contains(url)
    else:
        assert not selenium_generics.does_current_url_contains(url)


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The url '(?P<url>.*)' is opened in a new (tab|window)"),
    converters=dict(url=env_variables.env_formatted), )
def check_is_opened_in_new_window(selenium_generics: SeleniumGenerics, soft_assert: str, url: str):
    selenium_generics.switch_to_last_window()
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.does_url_equals(url)
    else:
        assert selenium_generics.does_url_equals(url)


# WEB context Predefined Step
@given(parsers.re("I refresh the current page"))
@when(parsers.re("I refresh the current page"))
def refresh_page(selenium_generics: SeleniumGenerics):
    selenium_generics.refresh_page()


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

HTML

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@given(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
def check_property_is(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str, attribute, wait_for):
    wait_for = int(wait_for) if wait_for else None
    if wait_for:
        try:
            polling.poll(lambda: selenium_generics.get_attribute_of_element(locators.parse_and_get(locator_path, selenium_generics), attribute) == value, step=2, timeout=wait_for)
        except (TimeoutException, polling.TimeoutException):
            pass
    actual_value = selenium_generics.get_attribute_of_element(
        locators.parse_and_get(locator_path, selenium_generics), attribute)
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_value).is_equal_to(value)
    else:
        assert_that(actual_value).is_equal_to(value)


# WEB context Predefined Step
@given(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is not the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is not the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is not the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
def check_property_is_not(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str, attribute, wait_for):
    wait_for = int(wait_for) if wait_for else None
    if wait_for:
        try:
            polling.poll(lambda: selenium_generics.get_attribute_of_element(locators.parse_and_get(locator_path, selenium_generics),attribute) != value, step=2, timeout=wait_for)
        except (TimeoutException, polling.TimeoutException):
            pass
    actual_value = selenium_generics.get_attribute_of_element(
        locators.parse_and_get(locator_path, selenium_generics), attribute)
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_value).is_not_equal_to(value)
    else:
        assert_that(actual_value).is_not_equal_to(value)


# WEB context Predefined Step
@given(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The css attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The css attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The css attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
def check_css_property_is(attribute, selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str, wait_for):
    wait_for = int(wait_for) if wait_for else None
    if wait_for:
        try:
            polling.poll(lambda: selenium_generics.get_css_value(locators.parse_and_get(locator_path, selenium_generics), attribute) == value, step=2, timeout=wait_for)
        except (TimeoutException, polling.TimeoutException):
            pass
    actual_value = selenium_generics.get_css_value(locators.parse_and_get(locator_path, selenium_generics), attribute)
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_value).is_equal_to(value)
    else:
        assert_that(actual_value).is_equal_to(value)


# WEB context Predefined Step
@given(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The css attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is not the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The css attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is not the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re(r"(With soft assertion '(?P<soft_assert>.*)' )?The css attribute '(?P<attribute>.*)' of element '(?P<locator_path>.*)' is not the value '(?P<value>.*)'(\s+)?((?:within)\s+(?:')(?P<wait_for>\w+)(?:') seconds)?$"),
    converters=dict(value=env_variables.env_formatted), )
def check_css_property_is_not(attribute, selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str, wait_for):
    wait_for = int(wait_for) if wait_for else None
    if wait_for:
        try:
            polling.poll(lambda: selenium_generics.get_css_value(locators.parse_and_get(locator_path, selenium_generics), attribute) != value, step=2, timeout=wait_for)
        except (TimeoutException, polling.TimeoutException):
            pass
    actual_value = selenium_generics.get_css_value(locators.parse_and_get(locator_path, selenium_generics), attribute)
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_value).is_not_equal_to(value)
    else:
        assert_that(actual_value).is_not_equal_to(value)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

VISIBILITY & EXISTENCE

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?There is an element '(?P<locator_path>.*)' on the page"))
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is displayed"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?There is an element '(?P<locator_path>.*)' on the page"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is displayed"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is displayed"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' (is displayed|exists)"))
def element_displayed(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert selenium_generics.is_element_visible(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not visible."
            else:
                assert selenium_generics.is_element_visible(
                    locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not visible."
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert selenium_generics.is_element_visible(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not visible."
        else:
            assert selenium_generics.is_element_visible(
                locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not visible."


@when(parsers.re("I click item '(?P<inner_text>.*)' for element '(?P<locator_path>.*)'"),
    converters=dict(inner_text=env_variables.env_formatted), )
def add_item_for_element(selenium_generics: SeleniumGenerics, locators: Locators, inner_text: str, locator_path: str):
    locator = locators.parse_and_get(locator_path, selenium_generics)
    locator = locator.format(inner_text)
    selenium_generics.click(locator)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?There is no element '(?P<locator_path>.*)' on the page"))
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not displayed"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?There is no element '(?P<locator_path>.*)' on the page"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not displayed"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not displayed"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' (is not displayed|does not exist)"))
def element_not_displayed(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert selenium_generics.is_element_invisible(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is visible."
            else:
                assert selenium_generics.is_element_invisible(
                    locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is visible."
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert selenium_generics.is_element_invisible(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is visible."
        else:
            assert selenium_generics.is_element_invisible(
                locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is visible."


# WEB & MOBILE contexts Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' appears exactly '(?P<occurrence_count>.*)' times"),
    converters=dict(occurrence_count=env_variables.env_formatted), )
def check_element_exists(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, occurrence_count: int, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert (len(selenium_generics.get_elements(locators.parse_and_get(locator_path, selenium_generics))) == int(occurrence_count))
            else:
                assert (len(selenium_generics.get_elements(
                    locators.parse_and_get(locator_path, selenium_generics))) == int(occurrence_count))
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert (len(selenium_generics.get_elements(locators.parse_and_get(locator_path, selenium_generics))) == int(occurrence_count))
        else:
            assert (len(selenium_generics.get_elements(
                locators.parse_and_get(locator_path, selenium_generics))) == int(occurrence_count))


# WEB & MOBILE contexts Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' does not appear exactly '(?P<occurrence_count>.*)' times"),
    converters=dict(occurrence_count=env_variables.env_formatted), )
def check_element_not_exists(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, occurrence_count: int, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert (len(selenium_generics.get_elements(locators.parse_and_get(locator_path, selenium_generics))) != int(occurrence_count))
            else:
                assert (len(selenium_generics.get_elements(
                    locators.parse_and_get(locator_path, selenium_generics))) != int(occurrence_count))
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert (len(selenium_generics.get_elements(locators.parse_and_get(locator_path, selenium_generics))) != int(occurrence_count))
        else:
            assert (len(selenium_generics.get_elements(
                locators.parse_and_get(locator_path, selenium_generics))) != int(occurrence_count))


# WEB & MOBILE contexts Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' becomes visible within '(?P<wait_for>.*)' seconds"),
    converters=dict(wait_for=env_variables.env_formatted), )
def wait_for_displayed(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path: str, wait_for: int, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert selenium_generics.is_element_visible(locators.parse_and_get(locator_path, selenium_generics), int(wait_for))
            else:
                assert selenium_generics.is_element_visible(locators.parse_and_get(locator_path, selenium_generics), int(wait_for))
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert selenium_generics.is_element_visible(locators.parse_and_get(locator_path, selenium_generics), int(wait_for))
        else:
            assert selenium_generics.is_element_visible(locators.parse_and_get(locator_path, selenium_generics), int(wait_for))


# WEB & MOBILE contexts Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' becomes invisible within '(?P<wait_for>.*)' seconds"),
    converters=dict(wait_for=env_variables.env_formatted), )
def wait_for_not_displayed(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path: str, wait_for: int, ):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert selenium_generics.is_element_invisible(locators.parse_and_get(locator_path, selenium_generics), int(wait_for))
            else:
                assert selenium_generics.is_element_invisible(locators.parse_and_get(locator_path, selenium_generics),
                                                              int(wait_for))
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert selenium_generics.is_element_invisible(locators.parse_and_get(locator_path, selenium_generics), int(wait_for))
        else:
            assert selenium_generics.is_element_invisible(locators.parse_and_get(locator_path, selenium_generics),
                                                          int(wait_for))


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' is within the viewport"))
def check_within_viewport(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert:str, locator_path: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics._is_in_viewport(locators.parse_and_get(locator_path, selenium_generics))
    else:
        assert selenium_generics._is_in_viewport(locators.parse_and_get(locator_path, selenium_generics))


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path>.*)' is not within the viewport"))
def check_within_viewport(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert not selenium_generics._is_in_viewport(locators.parse_and_get(locator_path, selenium_generics))
    else:
        assert not selenium_generics._is_in_viewport(locators.parse_and_get(locator_path, selenium_generics))


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

FORM ELEMENTS

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is enabled"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is enabled"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is enabled"))
def element_enabled(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert selenium_generics.is_enabled(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not enabled."
            else:
                assert selenium_generics.is_enabled(
                    locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not enabled."
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert selenium_generics.is_enabled(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not enabled."
        else:
            assert selenium_generics.is_enabled(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not enabled."


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not enabled"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not enabled"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not enabled"))
def element_not_enabled(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    assert not selenium_generics.is_enabled(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is enabled."
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                assert not selenium_generics.is_enabled(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is enabled."


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is selected"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is selected"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is selected"))
def element_selected(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics.is_selected(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not selected."
    else:
        assert selenium_generics.is_selected(
            locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is not selected."


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not selected"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not selected"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' is not selected"))
def element_not_selected(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert not selenium_generics.is_selected(locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is selected."
    else:
        assert not selenium_generics.is_selected(
            locators.parse_and_get(locator_path, selenium_generics)), f"Element {locator_path} is selected."


# WEB context Predefined Step
@given(parsers.re("I select the option at index '(?P<index>.*)' element '(?P<locator_path>.*)'"),
    converters=dict(index=env_variables.env_formatted), )
@when(parsers.re("I select the option at index '(?P<index>.*)' element '(?P<locator_path>.*)'"),
    converters=dict(index=env_variables.env_formatted), )
def select_option_by_index(selenium_generics: SeleniumGenerics, index: int, locators: Locators, locator_path):
    selenium_generics.select_by_index(locators.parse_and_get(locator_path, selenium_generics), int(index))


# WEB context Predefined Step
@given(parsers.re("I select the option '(?P<option>.*)' by value for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
@when(parsers.re("I select the option '(?P<option>.*)' by value for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
def select_option_by_value(selenium_generics: SeleniumGenerics, option: str, locators: Locators, locator_path):
    selenium_generics.select_by_value(locators.parse_and_get(locator_path, selenium_generics), option)


# WEB context Predefined Step
@given(parsers.re("I select the option '(?P<option>.*)' by visible text for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
@when(parsers.re("I select the option '(?P<option>.*)' by visible text for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
def select_option_by_visible_text(selenium_generics: SeleniumGenerics, option: str, locators: Locators, locator_path):
    selenium_generics.select_by_visible_text(locators.parse_and_get(locator_path, selenium_generics), option)


# WEB context Predefined Step
@given(parsers.re("I deselect the option at index '(?P<index>.*)' element '(?P<locator_path>.*)'"),
    converters=dict(index=env_variables.env_formatted), )
@when(parsers.re("I deselect the option at index '(?P<index>.*)' element '(?P<locator_path>.*)'"),
    converters=dict(index=env_variables.env_formatted), )
def deselect_option_index(selenium_generics: SeleniumGenerics, index: int, locators: Locators, locator_path):
    selenium_generics.deselect_by_index(locators.parse_and_get(locator_path, selenium_generics), int(index))


# WEB context Predefined Step
@given(parsers.re("I deselect the option '(?P<option>.*)' by value for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
@when(parsers.re("I deselect the option '(?P<option>.*)' by value for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
def deselect_option_by_value(selenium_generics: SeleniumGenerics, option: str, locators: Locators, locator_path: str, ):
    selenium_generics.deselect_by_value(locators.parse_and_get(locator_path, selenium_generics), option)


# WEB context Predefined Step
@given(parsers.re("I deselect the option '(?P<option>.*)' by visible text for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
@when(parsers.re("I deselect the option '(?P<option>.*)' by visible text for element '(?P<locator_path>.*)'"),
    converters=dict(option=env_variables.env_formatted), )
def deselect_option_by_visible_text(selenium_generics: SeleniumGenerics, option: str, locators: Locators,
                                    locator_path: str, ):
    selenium_generics.deselect_by_visible_text(locators.parse_and_get(locator_path, selenium_generics), option)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("I select the value '(?P<value>.*)' from dropdown '(?P<locator_path>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("I select the value '(?P<value>.*)' from dropdown '(?P<locator_path>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def select_dropdown_by_value(selenium_generics: SeleniumGenerics, value: str, locators: Locators, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.select_dropdown_value(locators.parse_and_get(locator_path, selenium_generics), value)
    else:
        selenium_generics.select_dropdown_value(locators.parse_and_get(locator_path, selenium_generics), value)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("I select the value at index '(?P<index>.*)' from dropdown '(?P<locator_path>.*)'"),
    converters=dict(index=env_variables.env_formatted), )
@when(parsers.re("I select the value at index '(?P<index>.*)' from dropdown '(?P<locator_path>.*)'"),
    converters=dict(index=env_variables.env_formatted), )
def select_dropdown_by_index(selenium_generics: SeleniumGenerics, index: int, locators: Locators, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.select_dropdown_value_at_index(locators.parse_and_get(locator_path, selenium_generics),
                                                             int(index))
    else:
        selenium_generics.select_dropdown_value_at_index(locators.parse_and_get(locator_path, selenium_generics),
                                                         int(index))


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect the selected dropdown '(?P<locator_path>.*)' text is '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect the selected dropdown '(?P<locator_path>.*)' text is '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I expect the selected dropdown '(?P<locator_path>.*)' text is '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def get_selected_dropdown_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str):
    actual_text = selenium_generics.get_first_selected_option(
        locators.parse_and_get(locator_path, selenium_generics)).text
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).is_equal_to(value)
    else:
        assert_that(actual_text).is_equal_to(value)


@then(parsers.re("I expect that drop-down list '(?P<locator_path>.*)' has in that specific order, only the values:(?P<table_values>.*)",flags=re.S),
    converters=dict(table_values=data_table_vertical_converter))
def check_dropdown_equal_values(selenium_generics, locators, locator_path, table_values):
    selenium_generics.compare_expected_and_ui_values_with_order(
        locators.parse_and_get(locator_path, selenium_generics), table_values)


@then(parsers.re("I expect that drop-down list '(?P<locator_path>.*)' contains the values:(?P<table_values>.*)",flags=re.S),
    converters=dict(table_values=data_table_vertical_converter))
def check_dropdown_contains_values(selenium_generics, locators, locator_path, table_values):
    selenium_generics.contains_expected_and_ui_values(
        locators.parse_and_get(locator_path, selenium_generics), table_values)


@then(parsers.re("I expect that drop-down list '(?P<locator_path>.*)' does not contains the values:(?P<table_values>.*)",flags=re.S),
    converters=dict(table_values=data_table_vertical_converter))
def check_dropdown_not_contains_values(selenium_generics, locators, locator_path, table_values):
    selenium_generics.does_not_contains_expected_and_ui_values(
        locators.parse_and_get(locator_path, selenium_generics), table_values)


# WEB context Predefined Step
@given(parsers.re("I select an item index '(?P<index>.*)' from a searchable dropdown with '(?P<type>label|placeholder|visible text)' '(?P<label>.*)' and search text '(?P<text>.*)'(\s+)?((?:with wait time)\s+(?:')(?P<timeout>.*)(?:') seconds)?$"),
      converters=dict(label=env_variables.env_formatted, text=env_variables.env_formatted))
@when(parsers.re("I select an item index '(?P<index>.*)' from a searchable dropdown with '(?P<type>label|placeholder|visible text)' '(?P<label>.*)' and search text '(?P<text>.*)'(\s+)?((?:with wait time)\s+(?:')(?P<timeout>.*)(?:') seconds)?$"),
      converters=dict(label=env_variables.env_formatted, text=env_variables.env_formatted))
def select_item_from_searchable_dropdown(selenium_generics: SeleniumGenerics, locators, index: str, label: str, text: str, timeout: str, type: str):
    def _select_element_from_dropdown_by_index(idx, dropdown_element_locator):
        num_of_elements = len(selenium_generics.get_elements(locator=dropdown_element_locator))
        if num_of_elements == 1:
            selenium_generics.enter_text(locator=locator, text_to_enter=text)
            time.sleep(int(timeout))
            for _ in range(int(idx)):
                selenium_generics.press_key("ARROW_DOWN")
                time.sleep(1)
            selenium_generics.press_key("RETURN")
        else:
            raise NoSuchElementException(f"Element cannot be uniquely identified. Found: {num_of_elements} elements")

    timeout = timeout if timeout else 10  # timeout until searchable dropdown appears
    if type == 'visible text':
        locator = locators.get_searchable_dropdown_by_visible_text(label)
        _select_element_from_dropdown_by_index(idx=index, dropdown_element_locator=locator)
    elif type == 'placeholder':
        locator = locators.get_element_by_attribute(type, label)
        _select_element_from_dropdown_by_index(idx=index, dropdown_element_locator=locator)
    elif type == 'label':
        locator = locators.get_searchable_dropdown_by_visible_text(value=label, attr="label")
        _select_element_from_dropdown_by_index(idx=index, dropdown_element_locator=locator)
    else:
        raise ValueError(f"Provide either 'placeholder', 'label' or 'visible text. Provided selector type: {type}")


# WEB context Predefined Step
@given(parsers.re(r"I select by visible text (?:')(?P<value>.*)(?:') from dropdown (whose placeholder is equal to (?:')(?P<placeholder>.*)(?:'))?(\s+)?((?:with the parent)\s+(?:')(?P<label>.*)(?:') label)?$"),
       converters=dict(value=env_variables.env_formatted,))
@when(parsers.re(r"I select by visible text (?:')(?P<value>.*)(?:') from dropdown (whose placeholder is equal to (?:')(?P<placeholder>.*)(?:'))?(\s+)?((?:with the parent)\s+(?:')(?P<label>.*)(?:') label)?$"),
       converters=dict(value=env_variables.env_formatted,))
def select_by_visible_text_from_dropdown(selenium_generics: SeleniumGenerics, locators, value: str, placeholder: str, label: str):
    if placeholder:
        locator = locators.get_dropdown_option(placeholder)
        num_of_elements = len(selenium_generics.get_elements(locator))
        if num_of_elements == 1:
            selenium_generics.select_by_visible_text(select_locator=locator, text=value)
        else:
            raise NoSuchElementException(f"Element cannot be uniquely identified. Found: {num_of_elements} elements")
    elif label:
        locator = locators.get_select_element_by_parent_label(label)
        num_of_elements = len(selenium_generics.get_elements(locator))
        if num_of_elements == 1:
            selenium_generics.select_by_visible_text(select_locator=locator, text=value)
        else:
            raise NoSuchElementException(f"Element cannot be uniquely identified. Found: {num_of_elements} elements")
    else:
        raise ValueError(f"Provide either 'placeholder' or 'label'. Provided placeholder: {placeholder}; label: {label}")



"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

TEXT

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@given(parsers.re("I add text '(?P<value>.*)' to field '(?P<locator_path>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("I add text '(?P<value>.*)' to field '(?P<locator_path>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def add_element_value(selenium_generics: SeleniumGenerics, value: str, locators: Locators, locator_path: str):
    current_text = selenium_generics.get_input_text(locators.parse_and_get(locator_path, selenium_generics))
    selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), f"{current_text}{value}")


# WEB context Predefined Step
@given(parsers.re("I add text '(?P<text>.*)' in input field whose attribute '(?P<attribute>.*)' is equal to '(?P<value>.*)'"),
    converters=dict(text=env_variables.env_formatted), )
@when(parsers.re("I add text '(?P<text>.*)' in input field whose attribute '(?P<attribute>.*)' is equal to '(?P<value>.*)'"),
    converters=dict(text=env_variables.env_formatted), )
def add_text_based_on_attribute(selenium_generics: SeleniumGenerics, locators, text: str, attribute: str, value: str):
    locator = locators.get_element_by_attribute(attribute, value)
    num_of_elements = len(selenium_generics.get_elements(locator))
    if num_of_elements == 1:
        selenium_generics.clear_text(locator)
        selenium_generics.enter_text(locator, text)
    else:
        raise NoSuchElementException(f"Element cannot be uniquely identified. Found: {num_of_elements} elements")


# WEB context Predefined Step
@given(parsers.re("I attach file '(?P<file_path>.*)' to input field '(?P<locator_path>.*)'"),
    converters=dict(file_path=env_variables.env_formatted), )
@when(parsers.re("I attach file '(?P<file_path>.*)' to input field '(?P<locator_path>.*)'"),
    converters=dict(file_path=env_variables.env_formatted), )
def attach_file(selenium_generics: SeleniumGenerics, file_path: str, locators: Locators, locator_path):
    relative_path = Path(file_path)
    if relative_path and relative_path.absolute().is_file():
        full_path = relative_path.absolute().as_posix()
        selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), full_path)
    else:
        raise FileNotFoundError(f"The file {file_path}' is not a valid path to a file")


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("I set text '(?P<value>.*)' to field '(?P<locator_path>.*)'"),
       converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("I set text '(?P<value>.*)' to field '(?P<locator_path>.*)'"),
      converters=dict(value=env_variables.env_formatted), )
def set_element_value(selenium_generics: SeleniumGenerics, value: str, locators: Locators, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), value)
    else:
        selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), value)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re(r"I add current date to '(?P<locator_path>.*)' with '(?P<date_format>MM/dd/yyyy|MM/dd/yy|dd/MM/yyyy|dd/MM/yy|dd MMM yyyy)'"))
@when(parsers.re(r"I add current date to '(?P<locator_path>.*)' with '(?P<date_format>MM/dd/yyyy|MM/dd/yy|dd/MM/yyyy|dd/MM/yy|dd MMM yyyy)'"))
def add_current_date_for_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, date_format: str):
    if date_format == "MM/dd/yyyy":
        date_text = datetime.now().date().strftime("%m/%d/%Y")
    elif date_format == "MM/dd/yy":
        date_text = datetime.now().date().strftime("%m/%d/%y")
    elif date_format == "dd/MM/yyyy":
        date_text = datetime.now().date().strftime("%d/%m/%Y")
    elif date_format == "dd/MM/yy":
        date_text = datetime.now().date().strftime("%d/%m/%y")
    elif date_format == "dd MMM yyyy":
        date_text = datetime.now().date().strftime("%d %b %Y")
    else:
        raise ValueError(f"Date format: {date_format} is invalid")
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), date_text)
    else:
        selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), date_text)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re(r"I add '(?P<direction>Past|Current|Future)' time to '(?P<locator_path>.*)' with '(?P<time_format>HH:MM:SS|HH:MM)' format(\s+)?((?:and clock format)\s+(?:')(?P<clock_format>\w+)(?:'))?(\s+)?((?:and delimiter)\s+(?:')(?P<delimiter>.*)(?:'))?$"))
@when(parsers.re(r"I add '(?P<direction>Past|Current|Future)' time to '(?P<locator_path>.*)' with '(?P<time_format>HH:MM:SS|HH:MM)' format(\s+)?((?:and clock format)\s+(?:')(?P<clock_format>\w+)(?:'))?(\s+)?((?:and delimiter)\s+(?:')(?P<delimiter>.*)(?:'))?$"))
def add_custom_time_for_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, time_format: str, direction: str, delimiter: str, clock_format: str):
    delimiter = delimiter if delimiter else ":"
    clock_format = clock_format if clock_format else "24h"
    now = datetime.now()
    _from = now - now.replace(hour=0, minute=0, second=0, microsecond=0)
    _to = now.replace(hour=23, minute=59, second=59, microsecond=0) - now
    if clock_format == "24h":
        cf = "%H"
    elif clock_format == "12h":
        cf = "%I"
    else:
        raise ValueError(f"Clock format: {clock_format} is invalid")

    if direction == "Future":
        random_seconds = random.randint(1, _to.seconds)
        delta = timedelta(seconds=random_seconds)
        future_date = now + delta
        _time = future_date
    elif direction == "Past":
        random_seconds = random.randint(1, _from.seconds)
        delta = timedelta(seconds=random_seconds)
        past_date = now - delta
        _time = past_date
    elif direction == "Current":
        _time = now
    else:
        raise ValueError(f"Time direction: {direction} is invalid")

    if time_format == "HH:MM:SS":
        time_text = _time.strftime(f"{cf}{delimiter}%M{delimiter}%S")
    elif time_format == "HH:MM":
        time_text = _time.strftime(f"{cf}{delimiter}%M")
    else:
        raise ValueError(f"Time format: {time_format} is invalid")

    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), time_text)
    else:
        selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), time_text)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re(r"I add random '(?P<direction>future|past)' date to '(?P<locator_path>.*)' with '(?P<date_format>MM/dd/yyyy|MM/dd/yy|dd/MM/yyyy|dd/MM/yy|dd MMM yyyy)' format"))
@when(parsers.re(r"I add random '(?P<direction>future|past)' date to '(?P<locator_path>.*)' with '(?P<date_format>MM/dd/yyyy|MM/dd/yy|dd/MM/yyyy|dd/MM/yy|dd MMM yyyy)' format"))
def add_custom_date_for_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, date_format: str, direction: str):
    now = datetime.now()
    delta = timedelta(days=random.randint(1, 365 * 20))
    if direction == "future":
        random_date = now + delta
    elif direction == "past":
        random_date = now - delta
    else:
        raise ValueError(f"Time direction: {direction} is invalid")

    if date_format == "MM/dd/yyyy":
        date_text = random_date.strftime("%m/%d/%Y")
    elif date_format == "MM/dd/yy":
        date_text = random_date.strftime("%m/%d/%y")
    elif date_format == "dd/MM/yyyy":
        date_text = random_date.strftime("%d/%m/%Y")
    elif date_format == "dd/MM/yy":
        date_text = random_date.strftime("%d/%m/%y")
    elif date_format == "dd MMM yyyy":
        date_text = random_date.strftime("%d %b %Y")
    else:
        raise ValueError(f"Date format: {date_format} is invalid")
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), date_text)
    else:
        selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), date_text)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re(r"I add random string of length (?:')(?P<length>.*)(?:') composed of (?:')(?P<character_type>alphabetic characters|numeric characters|alphabetic and numeric characters)(?:') to field (?:')(?P<locator_path>.*)(?:')(\s+)?((?:and save as)\s+(?:')(?P<storage_var>\w+)(?:') environment variable)?$"))
@when(parsers.re(r"I add random string of length (?:')(?P<length>.*)(?:') composed of (?:')(?P<character_type>alphabetic characters|numeric characters|alphabetic and numeric characters)(?:') to field (?:')(?P<locator_path>.*)(?:')(\s+)?((?:and save as)\s+(?:')(?P<storage_var>\w+)(?:') environment variable)?$"))
def set_random_element_value(selenium_generics: SeleniumGenerics, length: str, character_type: str, locators: Locators, locator_path, storage_var):
    storage_var = storage_var if storage_var else None
    if length.isdigit():
        random_string = DataUtils.get_random_text(length=int(length), source=character_type)
        if MOBILE_SUFFIX in locator_path:
            with context_manager(selenium_generics):
                selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), random_string)
        else:
            selenium_generics.enter_text(locators.parse_and_get(locator_path, selenium_generics), random_string)
        if storage_var:
            os.environ[storage_var] = random_string
    else:
        raise ValueError(f"Wrong length value: {length}")


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("I clear text from field '(?P<locator_path>.*)'"))
@when(parsers.re("I clear text from field '(?P<locator_path>.*)'"))
def clear_text(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.clear_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        selenium_generics.clear_text(locators.parse_and_get(locator_path, selenium_generics))


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' text is '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' text is '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' text is '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def element_equals_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).is_equal_to(value)
    else:
        assert_that(actual_text).is_equal_to(value)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' text is '(?P<comparison_value>EQUAL|LESS_THAN|LESS_THAN_OR_EQUAL|GREATER_THAN|GREATER_THAN_OR_EQUAL)' to '(?P<value>.*)'"),
      converters=dict(value=env_variables.env_formatted))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' text is '(?P<comparison_value>EQUAL|LESS_THAN|LESS_THAN_OR_EQUAL|GREATER_THAN|GREATER_THAN_OR_EQUAL)' to '(?P<value>.*)'"),
      converters=dict(value=env_variables.env_formatted))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The element '(?P<locator_path>.*)' text is '(?P<comparison_value>EQUAL|LESS_THAN|LESS_THAN_OR_EQUAL|GREATER_THAN|GREATER_THAN_OR_EQUAL)' to '(?P<value>.*)'"),
      converters=dict(value=env_variables.env_formatted))
def compare_numbers(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str, soft_assert: str, comparison_value: str, value: str):
    comparison_values = dict(EQUAL=lambda number_1, number_2: assert_that(number_1).is_equal_to(number_2),
                             LESS_THAN=lambda number_1, number_2: assert_that(number_1).is_less_than(number_2),
                             LESS_THAN_OR_EQUAL=lambda number_1, number_2: assert_that(number_1).is_less_than_or_equal_to(number_2),
                             GREATER_THAN=lambda number_1, number_2: assert_that(number_1).is_greater_than(number_2),
                             GREATER_THAN_OR_EQUAL=lambda number_1, number_2: assert_that(number_1).is_greater_than_or_equal_to(number_2))
    number_value = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))

    if number_value.isnumeric() and number_value.strip().isdigit():
        number_value = int(number_value.strip())
    else:
        raise ValueError(f"Given number_value: {number_value} is not in a numeric integer value")

    compare_function = comparison_values[comparison_value]

    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            if soft_assert is not None and soft_assert.lower() == 'true':
                with check:
                    compare_function(number_value, int(value))
            else:
                compare_function(number_value, int(value))
    else:
        if soft_assert is not None and soft_assert.lower() == 'true':
            with check:
                compare_function(number_value, int(value))
        else:
            compare_function(number_value, int(value))


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' contains the text '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' contains the text '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' contains the text '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def contains_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).contains(value)
    else:
        assert_that(actual_text).contains(value)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' contains any text"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' contains any text"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' contains any text"))
def contains_any_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).matches(r"(.*?)")
    else:
        assert_that(actual_text).matches(r"(.*?)")


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' text is not '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' text is not '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' text is not '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def element_not_equals_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).is_not_equal_to(value)
    else:
        assert_that(actual_text).is_not_equal_to(value)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' does not contain the text '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' does not contain the text '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' does not contain the text '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def does_not_contain_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).does_not_contain(value)
    else:
        assert_that(actual_text).does_not_contain(value)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' does not contain any text"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' does not contain any text"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' does not contain any text"))
def does_not_contain_any_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).matches(r"^$")
    else:
        assert_that(actual_text).matches(r"^$")


# WEB context Predefined Step
@given(parsers.re("I clear text using keys from field '(?P<locator_path>.*)'"))
@when(parsers.re("I clear text using keys from field '(?P<locator_path>.*)'"))
def clear_text_actions(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    selenium_generics.clear_text_using_keys(locators.parse_and_get(locator_path, selenium_generics))


# WEB & MOBILE contexts Predefined Step
@then(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path1>.*)' contains the same text as element '(?P<locator_path2>.*)'$"))
def check_contains_same_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path1, locator_path2, ):
    if MOBILE_SUFFIX in locator_path1 and MOBILE_SUFFIX in locator_path2:
        with context_manager(selenium_generics):
            actual_text1 = selenium_generics.get_element_text(locators.parse_and_get(locator_path1, selenium_generics))
            actual_text2 = selenium_generics.get_element_text(locators.parse_and_get(locator_path2, selenium_generics))
    else:
        actual_text1 = selenium_generics.get_element_text(locators.parse_and_get(locator_path1, selenium_generics))
        actual_text2 = selenium_generics.get_element_text(locators.parse_and_get(locator_path2, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text1).is_equal_to(actual_text2)
    else:
        assert_that(actual_text1).is_equal_to(actual_text2)


# WEB & MOBILE contexts Predefined Step
@then(parsers.re(
    "(With soft assertion '(?P<soft_assert>.*)' )?I expect that element '(?P<locator_path1>.*)' does not contain the same text as element '(?P<locator_path2>.*)'$"))
def check_does_not_contain_same_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path1, locator_path2, ):
    if MOBILE_SUFFIX in locator_path1 and MOBILE_SUFFIX in locator_path2:
        with context_manager(selenium_generics):
            actual_text1 = selenium_generics.get_element_text(locators.parse_and_get(locator_path1, selenium_generics))
            actual_text2 = selenium_generics.get_element_text(locators.parse_and_get(locator_path2, selenium_generics))
    else:
        actual_text1 = selenium_generics.get_element_text(locators.parse_and_get(locator_path1, selenium_generics))
        actual_text2 = selenium_generics.get_element_text(locators.parse_and_get(locator_path2, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text1).is_not_equal_to(actual_text2)
    else:
        assert_that(actual_text1).is_not_equal_to(actual_text2)


# WEB & MOBILE contexts Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The trimmed text on (button|element) '(?P<locator_path>.*)' is '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def check_element_equals_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text.strip()).is_equal_to(value)
    else:
        assert_that(actual_text.strip()).is_equal_to(value)


# WEB & MOBILE contexts Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The trimmed text on (button|element) '(?P<locator_path>.*)' is not '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def check_element_not_equals_text(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path, value: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text.strip()).is_not_equal_to(value)
    else:
        assert_that(actual_text.strip()).is_not_equal_to(value)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

CONFIRMATION POP-UP

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A (alertbox|confirmbox|prompt) is opened"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A (alertbox|confirmbox|prompt) is opened"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A (alertbox|confirmbox|prompt) is opened"))
def step_presence_of_alert(selenium_generics: SeleniumGenerics, soft_assert: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert selenium_generics._wait_for_presence_of_an_alert()
    else:
        assert selenium_generics._wait_for_presence_of_an_alert()


# WEB context Predefined Step
@given(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A (alertbox|confirmbox|prompt) is not opened"))
@when(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A (alertbox|confirmbox|prompt) is not opened"))
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?A (alertbox|confirmbox|prompt) is not opened"))
def check_modal_not_present(selenium_generics: SeleniumGenerics, soft_assert: str):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert not selenium_generics._wait_for_presence_of_an_alert()
    else:
        assert not selenium_generics._wait_for_presence_of_an_alert()


# WEB context Predefined Step
@given(parsers.re("I accept popup prompt"))
@given(parsers.re("I accept popup alertbox"))
@given(parsers.re("I accept popup confirmbox"))
@when(parsers.re("I accept popup prompt"))
@when(parsers.re("I accept popup alertbox"))
@when(parsers.re("I accept popup confirmbox"))
def accept_alert(selenium_generics: SeleniumGenerics):
    selenium_generics.accept_alert()


# WEB context Predefined Step
@given(parsers.re("I dismiss popup prompt"))
@given(parsers.re("I dismiss popup alertbox"))
@given(parsers.re("I dismiss popup confirmbox"))
@when(parsers.re("I dismiss popup prompt"))
@when(parsers.re("I dismiss popup alertbox"))
@when(parsers.re("I dismiss popup confirmbox"))
def dismiss_modal(selenium_generics: SeleniumGenerics):
    selenium_generics.dismiss_alert()


# WEB context Predefined Step
@given(parsers.re("I enter (?P<text>.*) into popup alertbox"), converters=dict(text=env_variables.env_formatted), )
@given(parsers.re("I enter (?P<text>.*) into popup confirmbox"), converters=dict(text=env_variables.env_formatted), )
@given(parsers.re("I enter (?P<text>.*) into popup prompt"), converters=dict(text=env_variables.env_formatted), )
@when(parsers.re("I enter (?P<text>.*) into popup alertbox"), converters=dict(text=env_variables.env_formatted), )
@when(parsers.re("I enter (?P<text>.*) into popup confirmbox"), converters=dict(text=env_variables.env_formatted), )
@when(parsers.re("I enter (?P<text>.*) into popup prompt"), converters=dict(text=env_variables.env_formatted), )
def check_modal(selenium_generics: SeleniumGenerics, text: str):
    selenium_generics.answer_alert_prompt(text)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

MOUSE & TOUCH

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step - Chrome Only
@given(parsers.re(
    "(?P<wait_str>I wait for maximum (?P<wait_seconds>\\d+) seconds, and )?I click on '(?P<locator_path>.+)'"))
@when(parsers.re(
    "(?P<wait_str>I wait for maximum (?P<wait_seconds>\\d+) seconds, and )?I click on '(?P<locator_path>.+)'"))
def click_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str, wait_seconds: int, ):
    if selenium_generics.driver.name.lower() in ("safari", "firefox"):
        wait_time = float(wait_seconds or selenium_generics.click_by_action.__kwdefaults__["max_wait_time"])
        selenium_generics.click(locators.parse_and_get(locator_path, selenium_generics), max_wait_time=wait_time)
    else:
        func = selenium_generics.click_by_action
        wait_time = float(wait_seconds or func.__kwdefaults__["max_wait_time"])
        func(locators.parse_and_get(locator_path, selenium_generics), max_wait_time=wait_time)


# WEB context Predefined Step
@given(parsers.re("I click on type:'(?P<element_type>.*)' element with text equal to '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("I click on type:'(?P<element_type>.*)' element with text equal to '(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def click_on_text(selenium_generics: SeleniumGenerics, element_type, value: str):
    selenium_generics.click(f"//{element_type}[normalize-space()='{value}']")


# WEB context Predefined Step
@given(parsers.re("I click on type:'(?P<element_type>.*)' that contains the text:'(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
@when(parsers.re("I click on type:'(?P<element_type>.*)' that contains the text:'(?P<value>.*)'"),
    converters=dict(value=env_variables.env_formatted), )
def click_on_partial_text(selenium_generics: SeleniumGenerics, element_type, value: str):
    selenium_generics.click(f"//{element_type}[contains(normalize-space(),'{value}')]")


# WEB context Predefined Step
@given(parsers.re("I (double click|doubleclick) on '(?P<locator_path>.*)'"))
@when(parsers.re("I (double click|doubleclick) on '(?P<locator_path>.*)'"))
def dbl_click_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    selenium_generics.double_click(locators.parse_and_get(locator_path, selenium_generics))


# WEB context Predefined Step
@given(parsers.re("I select the radio option whose attribute '(?P<option_attribute>.*)' is equal to '(?P<text>.*)' from the parent attribute '(?P<parent_attribute>.*)' and value '(?P<value>.*)'"),
       converters=dict(text=env_variables.env_formatted), )
@when(parsers.re("I select the radio option whose attribute '(?P<option_attribute>.*)' is equal to '(?P<text>.*)' from the parent attribute '(?P<parent_attribute>.*)' and value '(?P<value>.*)'"),
       converters=dict(text=env_variables.env_formatted), )
def select_radio_option_from_parent_attribute(selenium_generics: SeleniumGenerics, locators, option_attribute: str, text: str, parent_attribute: str, value: str):
    locator = locators.get_radio_option_from_parent(option_attribute, text, parent_attribute, value)
    num_of_elements = len(selenium_generics.get_elements(locator))
    if num_of_elements == 1:
        selenium_generics.click(locator)
    else:
        raise NoSuchElementException(f"Element cannot be uniquely identified. Found: {num_of_elements} elements")


# WEB context Predefined Step
@given(parsers.re("I drag and drop element '(?P<source>.*)' to element '(?P<target>.*)'"))
@when(parsers.re("I drag and drop element '(?P<source>.*)' to element '(?P<target>.*)'"))
def drag_and_drop_element(selenium_generics: SeleniumGenerics, locators: Locators, source: str, target: str):
    selenium_generics.drag_and_drop(locators.parse_and_get(source, selenium_generics),
                                    locators.parse_and_get(target, selenium_generics))


# WEB context Predefined Step
@given(parsers.re("I drag and drop element '(?P<source>.*)' by offset '(?P<x>.*)' and '(?P<y>.*)'"),
       converters=dict(x=env_variables.env_formatted, y=env_variables.env_formatted), )
@when(parsers.re("I drag and drop element '(?P<source>.*)' by offset '(?P<x>.*)' and '(?P<y>.*)'"),
      converters=dict(x=env_variables.env_formatted, y=env_variables.env_formatted), )
def drag_and_drop_element_by_offset(selenium_generics: SeleniumGenerics, locators: Locators, source: str, x: int, y: int):
    selenium_generics.drag_and_drop_by_offset(locators.parse_and_get(source, selenium_generics), int(x), int(y))


# WEB context Predefined Step
@given(parsers.re("I scroll to element '(?P<locator_path>.*)'"))
@when(parsers.re("I scroll to element '(?P<locator_path>.*)'"))
def scroll_to_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    selenium_generics.is_element_in_viewport(locators.parse_and_get(locator_path, selenium_generics))
    # element is still not visible after scroll into view
    if not (selenium_generics.is_element_visible(locators.parse_and_get(locator_path, selenium_generics), 2)):
        raise NoSuchElementException(
            f"The web element {locator_path}' is not visible or accessible for interactions")


# WEB context Predefined Step
@given(parsers.re("I scroll to view and click on '(?P<locator_path>.*)'"))
@when(parsers.re("I scroll to view and click on '(?P<locator_path>.*)'"))
def click_on_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    locator = locators.parse_and_get(locator_path, selenium_generics)
    selenium_generics.scroll_into_view(locator)
    selenium_generics.click(locator)


# WEB context Predefined Step
@given(parsers.re("I hover over '(?P<locator_path>.*)'"))
@given(parsers.re("I move to element '(?P<locator_path>.*)'"))
@when(parsers.re("I hover over '(?P<locator_path>.*)'"))
@when(parsers.re("I move to element '(?P<locator_path>.*)'"))
def hover_over_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    locator = locators.parse_and_get(locator_path, selenium_generics)
    selenium_generics.scroll_into_view(locator)
    selenium_generics.hover(locator)


# WEB context Predefined Step
@given(parsers.re("I move to an element '(?P<locator_path>.*)' with offset '(?P<x>.*)' '(?P<y>.*)'"),
       converters=dict(x=env_variables.env_formatted, y=env_variables.env_formatted), )
@when(parsers.re("I move to an element '(?P<locator_path>.*)' with offset '(?P<x>.*)' '(?P<y>.*)'"),
      converters=dict(x=env_variables.env_formatted, y=env_variables.env_formatted), )
def move_to_element_by_offset(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, x: int, y: int):
    selenium_generics.move_to_element_by_offset(locators.parse_and_get(locator_path, selenium_generics), int(x), int(y))


# WEB context Predefined Step
@given(parsers.re("I hover over '(?P<locator_path1>.*)' and click element '(?P<locator_path2>.*)'"))
@when(parsers.re("I hover over '(?P<locator_path1>.*)' and click element '(?P<locator_path2>.*)'"))
def hover_over_and_click_sub_menu(
    selenium_generics: SeleniumGenerics,
    locators: Locators,
    locator_path1: str,
    locator_path2: str,
):
    main_menu = locators.parse_and_get(locator_path1, selenium_generics)
    sub_menu = locators.parse_and_get(locator_path2, selenium_generics)
    selenium_generics.scroll_into_view(main_menu)
    selenium_generics.hover_and_click(main_menu, sub_menu)


# WEB context Predefined Step
@given(parsers.re("I click on SVG element '(?P<locator_path>.*)'"))
@when(parsers.re("I click on SVG element '(?P<locator_path>.*)'"))
def click_svg_element(selenium_generics: SeleniumGenerics, locators: Locators, locator_path):
    selenium_generics.click_by_action(locators.parse_and_get(locator_path, selenium_generics))


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("I click on (checkbox|button|dropdown|item|element) '(?P<locator_path>.*)'"))
@given(parsers.re("I tap on '(?P<locator>.*)'"))
@when(parsers.re("I click on (checkbox|button|dropdown|item|element) '(?P<locator_path>.*)'"))
@when(parsers.re("I tap on '(?P<locator>.*)'"))
def click_on_locator(selenium_generics: SeleniumGenerics, locators: Locators, locator_path: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.click(locators.parse_and_get(locator_path, selenium_generics))
    else:
        selenium_generics.click(locators.parse_and_get(locator_path, selenium_generics))


# WEB contexts Predefined Step
@given(parsers.re("I click on element with visible text '(?P<visibility_option>EQUALS|CONTAINS|STARTS_WITH|ENDS_WITH)' '(?P<value>.*)'"),
       converters=dict(value=env_variables.env_formatted),)
@when(parsers.re("I click on element with visible text '(?P<visibility_option>EQUALS|CONTAINS|STARTS_WITH|ENDS_WITH)' '(?P<value>.*)'"),
       converters=dict(value=env_variables.env_formatted),)
def click_on_element_by_visible_text(selenium_generics: SeleniumGenerics, locators, visibility_option: str, value: str):
    locator = locators.get_element_by_text(value, visibility_option)
    num_of_elements = len(selenium_generics.get_elements(locator))
    if num_of_elements == 1:
        selenium_generics.click(locator)
    else:
        raise NoSuchElementException(f"Element cannot be uniquely identified. Found: {num_of_elements} elements")


# MOBILE contexts Predefined Step
@given(parsers.re("I tap '(?P<corner>BOTTOM_LEFT|BOTTOM_RIGHT|TOP_LEFT|TOP_RIGHT)' corner of element '(?P<locator_path>.*)'"))  # noqa
@when(parsers.re("I tap '(?P<corner>BOTTOM_LEFT|BOTTOM_RIGHT|TOP_LEFT|TOP_RIGHT)' corner of element '(?P<locator_path>.*)'"))  # noqa
def click_element_corner(selenium_generics: SeleniumGenerics, locators: Locators, corner: str, locator_path: str):
    """
    Taps the <corner> of the given element, in the given context
    """
    locator = locators.parse_and_get(locator_path, selenium_generics)
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            selenium_generics.tap_corner_of_element(selenium_generics, corner, locator)
    else:
        selenium_generics.tap_corner_of_element(selenium_generics, corner, locator)


# ANDROID MOBILE context Predefined Step
@given("On android, I tap on back navigation")
@when("On android, I tap on back navigation")
def tap_back_nav(selenium_generics: SeleniumGenerics):
    if selenium_generics.is_android():
        with context_manager(driver=selenium_generics):
            selenium_generics.back()


# iOS MOBILE context Predefined Step
@given(parsers.re("On iOS, I navigate back to app after clicking on '(?P<locator>.*)'"))
@when(parsers.re("On iOS, I navigate back to app after clicking on '(?P<locator>.*)'"))
def navigate_back_to_app(selenium_generics: SeleniumGenerics, locators: Locators, locator: str):
    if not selenium_generics.is_android():
        with context_manager(driver=selenium_generics):
            selenium_generics.click(locators.parse_and_get(locator, selenium_generics))


# MOBILE contexts Predefined Step
@given(parsers.re(r"I swipe '(?P<direction>left|right)' on element '(?P<locator_path>.*)(?:')(\s+)?((?:by)\s+(?:')(?P<pixels>\w+)(?:') px)?$"),
    converters=dict(pixels=env_variables.env_formatted), )
@when(parsers.re(r"I swipe '(?P<direction>left|right)' on element '(?P<locator_path>.*)(?:')(\s+)?((?:by)\s+(?:')(?P<pixels>\w+)(?:') px)?$"),
    converters=dict(pixels=env_variables.env_formatted), )
def swipe_horizontally_on_element(selenium_generics: SeleniumGenerics, locators: Locators, direction: str, locator_path, pixels: int):
    move_by_pixels = int(pixels) if pixels else 200
    with context_manager(driver=selenium_generics):
        element = selenium_generics.get_element(locators.parse_and_get(locator_path, selenium_generics))
        if direction == 'left':
            selenium_generics.swipe_left_on_element(element=element, pixels=move_by_pixels)
        elif direction == 'right':
            selenium_generics.swipe_right_on_element(element=element, pixels=move_by_pixels)


# MOBILE contexts Predefined Step
@given(parsers.re(r"I swipe '(?P<direction>left|right)' on the page(\s+)?((?:by)\s+(?:')(?P<pixels>\w+)(?:') px)?$"),
    converters=dict(pixels=env_variables.env_formatted), )
@when(parsers.re(r"I swipe '(?P<direction>left|right)' on the page(\s+)?((?:by)\s+(?:')(?P<pixels>\w+)(?:') px)?$"),
    converters=dict(pixels=env_variables.env_formatted), )
def swipe_to_the_next_page(selenium_generics: SeleniumGenerics, direction, pixels: int):
    move_by_pixels = int(pixels) if pixels else 700
    with context_manager(driver=selenium_generics):
        if direction == 'left':
            selenium_generics.swipe_left_by_coordinates(pixels=move_by_pixels)
        elif direction == 'right':
            selenium_generics.swipe_right_by_coordinates(pixels=move_by_pixels)


# MOBILE context Predefined Step
@given(parsers.re("I swipe down '(?P<percent>.*)' % each time for '(?P<number>.*)' times"),
    converters=dict(percent=env_variables.env_formatted, number=env_variables.env_formatted), )
@when(parsers.re("I swipe down '(?P<percent>.*)' % each time for '(?P<number>.*)' times"),
    converters=dict(percent=env_variables.env_formatted, number=env_variables.env_formatted), )
def swipe_down_each_time_in_percentage(selenium_generics: SeleniumGenerics, percent: int, number: int):
    with context_manager(driver=selenium_generics):
        selenium_generics.swipe_down(int(percent), int(number))


# MOBILE context Predefined Step
@given(parsers.re("I scroll to element '(?P<locator>.*)' for '(?P<iterations>.*)' iterations"),
    converters=dict(iterations=env_variables.env_formatted), )
@when(parsers.re("I scroll to element '(?P<locator>.*)' for '(?P<iterations>.*)' iterations"),
    converters=dict(iterations=env_variables.env_formatted), )
def scroll_to_native_element(selenium_generics: SeleniumGenerics, locators: Locators, locator: str,
                             iterations: int):
    for iterations in range(int(iterations)):
        if selenium_generics.is_element_visible(locators.parse_and_get(locator, selenium_generics)):
            break
        else:
            with context_manager(driver=selenium_generics):
                selenium_generics.swipe_to_element(selenium_generics)


# WEB & MOBILE contexts Predefined Step
@given(parsers.re("I scroll on the '(?P<scroll_direction>top|bottom)' of the page"))
@when(parsers.re("I scroll on the '(?P<scroll_direction>top|bottom)' of the page"))
def scroll_to_edge(selenium_generics: SeleniumGenerics, scroll_direction):
    """
    accepted directions: top,bottom
    """
    if hasattr(selenium_generics.driver, 'current_context'):
        with context_manager(driver=selenium_generics):
            selenium_generics.scroll_to_edge_of_mobile_page(selenium_generics, scroll_direction)
    else:
        selenium_generics.scroll_to_edge(scroll_direction)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

HTML TABLE

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


@then(parsers.re("I expect table '(?P<locator_path>.*)' headers ('(?P<header_tag_path>.*)' )?to match:(?P<data_table>.*)",
                 flags=re.S, ), converters=dict(data_table=data_table_horizontal_converter),)
def verify_table_headers_match_exactly(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, data_table, header_tag_path):
    header_tag = locators.parse_and_get(header_tag_path, selenium_generics) if header_tag_path else "//th"
    table_headers = selenium_generics.get_elements(f"{locators.parse_and_get(locator_path, selenium_generics)}{header_tag}")
    expected_columns = data_table[list(data_table.keys())[0]]
    columns_present = list()
    for table_header in table_headers:
            columns_present.append(table_header.text)
    assert_that(columns_present).is_equal_to(expected_columns)


@then(parsers.re("I expect table '(?P<locator_path>.*)' headers ('(?P<header_tag_path>.*)' )?to contain:(?P<data_table>.*)",
                 flags=re.S, ), converters=dict(data_table=data_table_horizontal_converter), )
def verify_table_headers_contain_columns(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, data_table, header_tag_path):
    header_tag = locators.parse_and_get(header_tag_path, selenium_generics) if header_tag_path else "//th"
    table_headers = selenium_generics.get_elements(f"{locators.parse_and_get(locator_path, selenium_generics)}{header_tag}")
    expected_columns = data_table[list(data_table.keys())[0]]
    columns_present = list()
    for table_header in table_headers:
            columns_present.append(table_header.text)
    assert_that(columns_present).contains(*expected_columns)


@then(parsers.re("I expect the column in table '(?P<locator_path>.*)' has the values:(?P<data_table>.*)",
                 flags=re.S, ), converters=dict(data_table=data_table_horizontal_converter), )
def verify_table_column_contain_values(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, data_table):
    table_locator = locators.parse_and_get(locator_path, selenium_generics)
    table_header_locator = f"{table_locator}//th | {table_locator}//th//*"
    table_headers = selenium_generics.get_elements(table_header_locator)
    column_names = list(data_table.keys())
    column_index_dict = dict()
    td_index = 0
    for table_header in table_headers:
        if table_header.tag_name == 'th':
            td_index += 1
        if table_header.text in column_names:
            column_index_dict[table_header.text] = td_index

    for column in column_names:
        for value in data_table[column]:
            cell_locator = f"{table_locator}//tr//td[{column_index_dict[column]}][text()='{value}'] | " \
                           f"{table_locator}//tr//td[{column_index_dict[column]}]//*[text()='{value}']"
            assert len(selenium_generics.get_elements(cell_locator)) > 0


@then(parsers.re("I expect that '(?P<row>.*)' row has the value '(?P<expected_text>.*)' in column '(?P<column>.*)' of table '(?P<locator_path>.*)'"),
    converters=dict(expected_text=env_variables.env_formatted), )
def verify_column_contain_value(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, row, column, expected_text: str):
    table_locator = locators.parse_and_get(locator_path, selenium_generics)
    table_header_locator = f"{table_locator}//th | {table_locator}//th//*"
    table_headers = selenium_generics.get_elements(table_header_locator)
    td_index = 0
    row_number = ''
    row_number = [row_number+i for i in row if i.isdigit()][0]
    for table_header in table_headers:
        if table_header.tag_name == 'th':
            td_index += 1
        if table_header.text == column:
            break

    cell_locator = f"{table_locator}//tr[{row_number}]//td[{td_index}][text()='{expected_text}'] | " \
                   f"{table_locator}//tr[{row_number}]//td[{td_index}]//*[text()='{expected_text}']"
    assert len(selenium_generics.get_elements(cell_locator)) > 0


@then(parsers.re("I expect that '(?P<row>.*)' row in table '(?P<locator_path>.*)' has the following values:(?P<data_table>.*)",
                 flags=re.S, ), converters=dict(data_table=data_table_horizontal_converter), )
def verify_table_row_contain_values(selenium_generics: SeleniumGenerics, locators: Locators, row, locator_path, data_table):
    table_locator = locators.parse_and_get(locator_path, selenium_generics)
    table_header_locator = f"{table_locator}//th | {table_locator}//th//*"
    table_headers = selenium_generics.get_elements(table_header_locator)
    column_names = list(data_table.keys())
    column_index_dict = dict()
    td_index = 0
    row_number = ''
    row_number = [row_number+i for i in row if i.isdigit()][0]
    for table_header in table_headers:
        if table_header.tag_name == 'th':
            td_index += 1
        if table_header.text in column_names:
            column_index_dict[table_header.text] = td_index

    if len(data_table[column_names[0]]) > 1:
        raise Exception(f"This step can only validate data in one specific row. Data Table from BDD has {len(data_table[column_names[0]]) } rows.")

    for column in column_names:
        cell_locator = f"{table_locator}//tr[{row_number}]//td[{column_index_dict[column]}][text()='{data_table[column][0]}'] | " \
                       f"{table_locator}//tr[{row_number}]//td[{column_index_dict[column]}]//*[text()='{data_table[column][0]}']"
        assert len(selenium_generics.get_elements(cell_locator)) > 0


@then(parsers.re("I expect that table '(?P<locator_path>.*)' has '(?P<value>.*)' rows"),
      converters=dict(value=env_variables.env_formatted), )
def verify_column_contain_value(selenium_generics: SeleniumGenerics, locators: Locators, locator_path, value: int):
    table_row_locator = f"{locators.parse_and_get(locator_path, selenium_generics)}//tr"
    assert_that(len(selenium_generics.get_elements(table_row_locator))).is_equal_to(value)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

KEYBOARD

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@given(parsers.re("I press '(?P<key>.*)', focused over '(?P<locator_path>.*)'"))
@when(parsers.re("I press '(?P<key>.*)', focused over '(?P<locator_path>.*)'"))
def press_key_on_element(selenium_generics: SeleniumGenerics, key: str, locators: Locators, locator_path: str):
    locator = locators.parse_and_get(locator_path, selenium_generics)
    logger.info("Pressing Key", locator=locator, key=key)
    selenium_generics.press_key_on_element(locator, key)


# WEB context Predefined Step
@given(parsers.re("I press '(?P<key>.*)'"))
@when(parsers.re("I press '(?P<key>.*)'"))
def press_key_not_focused_on_element(selenium_generics: SeleniumGenerics, key: str):
    logger.info("Pressing Key", key=key)
    selenium_generics.press_key(key)


# MOBILE context Predefined Step
@given(parsers.re("I long tap on element '(?P<locator>.*)'"))
@when(parsers.re("I long tap on element '(?P<locator>.*)'"))
def long_tap(selenium_generics: SeleniumGenerics, locators: Locators, locator: str):
    with context_manager(driver=selenium_generics):
        selenium_generics.long_tap(selenium_generics, locators.parse_and_get(locator, selenium_generics))


# WEB context Predefined Step - Chrome Only
@given(parsers.re(
    "(?P<gap_str>With gap of (?P<gap_seconds>\\d+) seconds, )?I type text '(?P<text>.+)' in the input field '(?P<locator_path>.+)'"),
    converters=dict(text=env_variables.env_formatted))
@when(parsers.re(
    "(?P<gap_str>With gap of (?P<gap_seconds>\\d+) seconds, )?I type text '(?P<text>.+)' in the input field '(?P<locator_path>.+)'"),
    converters=dict(text=env_variables.env_formatted))
def step_realistic_typing(selenium_generics: SeleniumGenerics, locators: Locators,
                          gap_seconds: typing.Union[float, None], text: str, locator_path: str, ):
    if selenium_generics.driver.name.lower() in ("safari", "firefox"):
        raise NotImplemented('This predefined step cannot be used with Safari or Firefox, please use '
                             'set_element_value() or add_element_value() instead')
    func = selenium_generics.simulate_realistic_typing
    gap_seconds = float(gap_seconds or func.__kwdefaults__["gap_seconds"])
    func(locators.parse_and_get(locator_path, selenium_generics), text, gap_seconds=gap_seconds)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

ANALYTICS

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@when(parsers.re("I pause for the next '(?P<waiting_time>.*)' seconds and then I get the Network logs"))
@then(parsers.re("I pause for the next '(?P<waiting_time>.*)' seconds and then I get the Network logs"))
def get_network_analytics(driver, selenium_generics, waiting_time: int):
    # we have to wait for the BE responses to be successfully received
    time.sleep(int(waiting_time))
    BPStorage.store_analytics(selenium_generics.retrieve_analytics_from_network_logs(
        [json.loads(lr["message"])["message"] for lr in driver.get_log("performance")]))


# WEB context Predefined Step
@when("I clear the Network logs")
@then("I clear the Network logs")
def clear_network_analytics(driver):
    # getting the performance logs will also clear the driver logs
    driver.get_log("performance")


# WEB context Predefined Step
@then(parsers.re("The Analytics payload from '(?P<analytics_path>.*)'(, merged with '(?P<optional_analytics_path>.*)',)? appears exactly '(?P<presence_no>.*)' times in the Network logs"))
def compare_network_analytics(driver, selenium_generics: SeleniumGenerics, analytics_path: str,
                              optional_analytics_path: str, presence_no: int):
    analytics = Analytics()
    analytics_payload = analytics.parse_and_get(analytics_path)
    if optional_analytics_path is not None:
        analytics_payload.update(analytics.parse_and_get(optional_analytics_path))
    selenium_generics.compare_analytics_logs(BPStorage.get_analytics(), analytics_payload, presence_no)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

VISUAL REGRESSION

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I verify images '(?P<name>.*)' have no visual regression"))
def image_visual_is_valid(soft_assert: str, name):
    """Step Definition to verify if two images are same (Standalone Visual Testing)

    Both Base Image and Test Image are saved in respective directories as defined by boilerplate
    framework, i.e. test_data/visualtesting/base and test_data/visualtesting/test directories, with the
    same name (argument name passed in feature file).

    Args:
        name: str - Image Name to perform visual regression.

    Asserts:
        If Base and Test Images are same. Else raises AssertionError.

    """
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert are_two_images_look_same(name)
    else:
        assert are_two_images_look_same(name)


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I verify that element '(?P<locator_path>.*)' is not visually regressed:(?P<data_table>.*)",
                 flags=re.S, ), converters=dict(data_table=data_table_horizontal_converter), )
def element_visual_is_valid(selenium_generics: SeleniumGenerics, locators: Locators, soft_assert: str, locator_path: str, data_table: dict):
    """Step Definition to verify if a particular webelement is not visually regressed.

    Base Image should be saved in output/screenshots/base directory (name as provided in
    the data table). Test Function would take screenshot of corresponding webelement as
    provided in locator path, and asserts if it is same as base image.

    Args:
        selenium_generics - SeleniumGenerics instance
        locators - Locators instance
        locator_path - as provided in feature file step.
        data_table - retrieved from feature file - scenario - step.

    Asserts:
        If the webelement captured during test is same as the base image provided.
    """
    locator = locators.parse_and_get(locator_path, selenium_generics)
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert are_two_webelements_look_same(data_table["base_image"][0], selenium_generics, locator)
    else:
        assert are_two_webelements_look_same(data_table["base_image"][0], selenium_generics, locator)


# WEB context Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?I verify the page is not visually regressed:(?P<data_table>.*)", flags=re.S),
      converters=dict(data_table=data_table_horizontal_converter), )
def page_visual_is_valid(selenium_generics: SeleniumGenerics, soft_assert: str, data_table: dict):
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert are_two_webpages_look_same(data_table["base_image"][0], selenium_generics)
    else:
        assert are_two_webpages_look_same(data_table["base_image"][0], selenium_generics)


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

API with BDD

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


# WEB context Predefined Step
@when(parsers.re("I '(?P<request_type>.*)' an API request having base URL '(?P<base_url>.*)' with endpoint '(?P<endpoint>.*)' with headers '(?P<headers>.*)' and payload '(?P<payload>.*)' and store the response value"))
def send_api_request_store_value_bp_storage(selenium_generics: SeleniumGenerics, request: FixtureRequest, request_type: str, base_url: str, endpoint: str, headers: str, payload: str):
    api_response_container = dict()
    set_request_endpoint(request, base_url=base_url, endpoint=endpoint, request_name=UI_API_CALL)
    set_request_headers(request, headers=headers, request_name=UI_API_CALL)
    add_json_payload(request, json_payload=payload, request_name=UI_API_CALL)
    make_api_request(request, api_response_container=api_response_container, request_type=request_type, request_name=UI_API_CALL)
    BPStorage.store_api_bdd_response(get_api_response(request, api_response_container=api_response_container, request_name=UI_API_CALL))


# WEB & MOBILE contexts Predefined Step
@then(parsers.re("(With soft assertion '(?P<soft_assert>.*)' )?The (button|element) '(?P<locator_path>.*)' has the text equal with the API BDD response key '(?P<key>.*)'"))
def api_contains_text(selenium_generics: SeleniumGenerics, locators, soft_assert: str, locator_path: str, key: str):
    if MOBILE_SUFFIX in locator_path:
        with context_manager(selenium_generics):
            actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    else:
        actual_text = selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))
    if soft_assert is not None and soft_assert.lower() == 'true':
        with check:
            assert_that(actual_text).is_equal_to(BPStorage.get_api_bdd_response().json()[key])
    else:
        assert_that(actual_text).is_equal_to(BPStorage.get_api_bdd_response().json()[key])


"""
----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------

EXCEL / CSV

----------------------------------------------------------------------------------------------------------------------
----------------------------------------------------------------------------------------------------------------------
"""


@when(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted, expected_text=env_variables.env_formatted))
@then(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted, expected_text=env_variables.env_formatted))
def cell_text_is_equal_to(cell:str, sheet_name: str, file_path: str, expected_text: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).is_equal_to(expected_text)


@when(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' contains '(?P<expected_text>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted, expected_text=env_variables.env_formatted))
@then(parsers.re("Text inside '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' contains '(?P<expected_text>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted, expected_text=env_variables.env_formatted))
def cell_text_contains(cell:str, sheet_name: str, file_path: str, expected_text: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).contains(expected_text)


@when(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' does not contain any text"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
@then(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' does not contain any text"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
def cell_text_is_empty(cell:str, sheet_name: str, file_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).is_none()


@when(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
@then(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
def cell_text_equals_element_text(selenium_generics, locators, cell:str, sheet_name: str, file_path: str, locator_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(sheet[cell].value).is_equal_to(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics)))


@when(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
@then(parsers.re("The '(?P<cell>.+)' on sheet '(?P<sheet_name>.+)' of excel file '(?P<file_path>.+)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
def element_text_contains_cell_text(selenium_generics, locators, cell:str, sheet_name: str, file_path: str, locator_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    assert_that(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))).contains(sheet[cell].value)


@given(parsers.re("I delete '(?P<file_path>.*)' file"))
@when(parsers.re("I delete '(?P<file_path>.*)' file"))
def delete_file(file_path: str):
    absolute_path = Path(file_path).absolute()
    if absolute_path.exists():
        os.remove(absolute_path)


@given(parsers.re("I create excel file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
@when(parsers.re("I create excel file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
def create_excel_file(file_name: str, file_path: str):
    if Path(file_name).suffix != '.xlsx':
        raise ValueError(f"Invalid file extension for {file_name}")
    excel_file = (Path(file_path) / file_name).absolute()
    openpyxl.Workbook().save(excel_file.as_posix())


@given(parsers.re("I write '(?P<text>.*)' to '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)'"),
       converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted,
       text=env_variables.env_formatted))
@when(parsers.re("I write '(?P<text>.*)' to '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted,
      text=env_variables.env_formatted))
def write_text_to_excel_file(text: str, cell: str, sheet_name: str, file_path: str):
    excel_file = Path(file_path).absolute().as_posix()
    wb = load_workbook(excel_file)
    wb[sheet_name][cell] = text
    wb.save(excel_file)


@given(parsers.re("I get text from '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variable with name '(?P<env_var>.*)'"),
       converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
@when(parsers.re("I get text from '(?P<cell>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variable with name '(?P<env_var>.*)'"),
      converters=dict(cell=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
def write_text_to_excel_file(cell: str, sheet_name: str, file_path: str, env_var: str):
    sheet = load_workbook(file_path)[sheet_name]
    text = sheet[cell].value
    os.environ[env_var] = text


@given(parsers.re("I get text of '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variables"),
       converters=dict(sheet_name=env_variables.env_formatted))
@when(parsers.re("I get text of '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' and save it as environment variables"),
      converters=dict(sheet_name=env_variables.env_formatted))
def store_excel_data_as_env_vars(sheet_name: str, file_path: str):
    sheet = load_workbook(file_path)[sheet_name]
    first_column = None
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(sheet.dimensions)
    for col in sheet.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row, values_only=False):
        if first_column:
            break
        for cell in col:
            if cell.value:
                first_column = cell.column
                break
    else:
        raise ValueError(f"No data found in {file_path}")
    for key, value in sheet.iter_rows(min_col=first_column, max_col=first_column + 1, min_row=min_row, max_row=max_row, values_only=True):
        if key and value:
            os.environ[key] = str(value)


@when(parsers.re("Number of total rows on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
@then(parsers.re("Number of total rows on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=env_variables.env_formatted, sheet_name=env_variables.env_formatted))
def total_rows_number_with_data_is_equal_to(sheet_name: str, file_path: str, row_count: str):
    sheet = load_workbook(file_path)[sheet_name]
    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@when(parsers.re("Number of rows containing '(?P<expected_text>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=env_variables.env_formatted, sheet_name=env_variables.env_formatted, text=env_variables.env_formatted))
@then(parsers.re("Number of rows containing '(?P<expected_text>.*)' on '(?P<sheet_name>.*)' of excel file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=env_variables.env_formatted, sheet_name=env_variables.env_formatted, text=env_variables.env_formatted))
def number_rows_with_text_is_equal_to(expected_text: str, sheet_name: str, file_path: str, row_count: str):
    sheet = load_workbook(file_path)[sheet_name]
    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value == f"{expected_text}" for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@when(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(expected_text=env_variables.env_formatted, cell=env_variables.env_formatted))
@then(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is equal to '(?P<expected_text>.+)'"),
      converters=dict(expected_text=env_variables.env_formatted, cell=env_variables.env_formatted))
def csv_cell_text_is_equal_to(cell: str, file_path: str, expected_text: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).is_equal_to(expected_text)


@when(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' contains '(?P<expected_text>.+)'"),
      converters=dict(expected_text=env_variables.env_formatted, cell=env_variables.env_formatted))
@then(parsers.re("Text inside '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' contains '(?P<expected_text>.+)'"),
      converters=dict(expected_text=env_variables.env_formatted, cell=env_variables.env_formatted))
def csv_cell_text_contains(cell: str, file_path: str, expected_text: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).contains(expected_text)


@when(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted))
@then(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' text is equal with the text of the '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted))
def csv_cell_text_equals_element_text(selenium_generics, locators, cell: str, file_path: str, locator_path: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).is_equal_to(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics)))


@when(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted))
@then(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' is contained in the text of '(?P<locator_path>.+)'"),
      converters=dict(cell=env_variables.env_formatted))
def element_text_contains_csv_cell_text(selenium_generics, locators, cell: str, file_path: str, locator_path: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(selenium_generics.get_element_text(locators.parse_and_get(locator_path, selenium_generics))).contains(
        sheet[cell].value)


@when(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' does not contain any text"),
      converters=dict(cell=env_variables.env_formatted))
@then(parsers.re("The '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' does not contain any text"),
      converters=dict(cell=env_variables.env_formatted))
def csv_cell_text_is_empty(cell: str, file_path: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
    assert_that(sheet[cell].value).is_empty()


@when(parsers.re("Number of rows containing '(?P<expected_text>.*)' of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(expected_text=env_variables.env_formatted, row_count=env_variables.env_formatted))
@then(parsers.re("Number of rows containing '(?P<expected_text>.*)' of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(expected_text=env_variables.env_formatted, row_count=env_variables.env_formatted))
def number_csv_rows_with_text_is_equal_to(expected_text: str, file_path: str, row_count: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)

    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value == f"{expected_text}" for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@when(parsers.re("Number of total rows of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=env_variables.env_formatted))
@then(parsers.re("Number of total rows of csv file '(?P<file_path>.*)' is '(?P<row_count>.*)'"),
      converters=dict(row_count=env_variables.env_formatted))
def total_csv_rows_number_with_data_is_equal_to(file_path: str, row_count: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)

    num_rows = 0
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            num_rows += 1
    assert_that(int(num_rows)).is_equal_to(int(row_count))


@given(parsers.re("I get text from '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' and save it as env variable with name '(?P<env_var>.*)'"),
       converters=dict(cell=env_variables.env_formatted))
@when(parsers.re("I get text from '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)' and save it as env variable with name '(?P<env_var>.*)'"),
       converters=dict(cell=env_variables.env_formatted))
def save_csv_cell_text_as_environment_variable(cell: str, file_path: str, env_var: str):
    csv_file = Path(file_path).absolute()
    if not all([csv_file.exists(), csv_file.is_file(), csv_file.suffix == '.csv']):
        raise FileNotFoundError(f"File {file_path} is not a valid csv file")
    wb = openpyxl.Workbook()
    sheet = wb.active
    with open(csv_file, newline="") as f:
        file = f.read()
    dialect = csv.Sniffer()
    delimiter = dialect.sniff(file).delimiter
    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)

    text = sheet[cell].value
    os.environ[env_var] = text


@given(parsers.re("I Write '(?P<text>.+)' to '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)'"),
       converters=dict(cell=env_variables.env_formatted, text=env_variables.env_formatted))
@when(parsers.re("I Write '(?P<text>.+)' to '(?P<cell>.*)' cell of csv file '(?P<file_path>.*)'"),
       converters=dict(cell=env_variables.env_formatted, text=env_variables.env_formatted))
def write_text_to_csv_cell(text: str, cell: str, file_path: str):
    csv_file = Path(file_path).absolute()
    wb = openpyxl.Workbook()
    sheet = wb.active
    delimiter = None

    if csv_file.exists():
        with open(csv_file, newline="") as f:
            file = f.read()
        if file:
            dialect = csv.Sniffer()
            delimiter = dialect.sniff(file).delimiter
    else:
        with open(csv_file, "w", newline="") as f:
            pass
    delimiter = delimiter if delimiter else ','

    with open(csv_file, newline="") as file:
        reader = csv.reader(file, delimiter=delimiter)
        for row in reader:
            sheet.append(row)
        sheet[cell] = text

    with open(csv_file, 'w', newline="") as f:
        writer = csv.writer(f, delimiter=delimiter)
        for row in sheet.iter_rows():
            writer.writerow([cell.value for cell in row])


@given(parsers.re("I create csv file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
@when(parsers.re("I create csv file '(?P<file_name>.*)' and save on '(?P<file_path>.+)'"))
def create_csv_file(file_name: str, file_path: str):
    if Path(file_name).suffix != '.csv':
        raise ValueError(f"Invalid file extension for {file_name}")
    csv_file = (Path(file_path) / file_name).absolute()

    # Create an empty csv file
    with open(csv_file, "w", newline="") as f:
        pass


"""
Sitemap step definitions have been removed in 3.11 and they can be found in custom.py in previous versions in case that recovery is needed.
"""
